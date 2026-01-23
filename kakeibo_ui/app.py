import uuid
import csv
import io
import base64
import sys
from pathlib import Path
from datetime import datetime

# プロジェクトルートをPythonパスに追加（ローカル実行時用）
project_root = Path(__file__).parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from flask import Flask, render_template, request, jsonify, Response
from shared.common.database.client import DatabaseClient

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Google Drive API
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    HAS_GOOGLE_DRIVE = True
except ImportError:
    HAS_GOOGLE_DRIVE = False

# パイプライン（レシート取り込み用）
try:
    from shared.pipeline import UnifiedDocumentPipeline
    from shared.common.connectors.google_drive import GoogleDriveConnector
    HAS_PIPELINE = True
except ImportError:
    HAS_PIPELINE = False
    UnifiedDocumentPipeline = None
    GoogleDriveConnector = None

app = Flask(__name__)

# Google Drive設定
GOOGLE_DRIVE_CREDENTIALS = None
_drive_service = None

def init_drive_service():
    """Google Drive APIサービスを初期化"""
    global _drive_service, GOOGLE_DRIVE_CREDENTIALS
    if not HAS_GOOGLE_DRIVE:
        return None
    if _drive_service:
        return _drive_service

    try:
        from shared.kakeibo.config import GOOGLE_DRIVE_CREDENTIALS as cred_path
        GOOGLE_DRIVE_CREDENTIALS = cred_path
        from pathlib import Path
        if Path(cred_path).exists():
            credentials = service_account.Credentials.from_service_account_file(
                cred_path,
                scopes=["https://www.googleapis.com/auth/drive.readonly"]
            )
            _drive_service = build("drive", "v3", credentials=credentials)
            return _drive_service
    except Exception as e:
        print(f"Google Drive初期化エラー: {e}")
    return None


def get_receipt_image_base64(drive_file_id: str) -> str:
    """Google Driveからレシート画像を取得してBase64エンコード"""
    service = init_drive_service()
    if not service or not drive_file_id:
        return None

    try:
        request = service.files().get_media(
            fileId=drive_file_id,
            supportsAllDrives=True
        )
        file_bytes = io.BytesIO()
        downloader = MediaIoBaseDownload(file_bytes, request)

        done = False
        while not done:
            status, done = downloader.next_chunk()

        file_bytes.seek(0)
        encoded = base64.b64encode(file_bytes.read()).decode('utf-8')
        return f"data:image/jpeg;base64,{encoded}"
    except Exception as e:
        print(f"画像取得エラー: {e}")
        return None


def get_db():
    """DB接続を取得"""
    return DatabaseClient(use_service_role=True).client


def check_auto_exclude(content, institution, rules):
    """自動除外ルールにマッチするかチェック（content + institution の組み合わせ）"""
    for rule in rules:
        content_pattern = rule.get('content_pattern', '')
        institution_pattern = rule.get('institution_pattern', '')
        if content_pattern in content and institution_pattern in institution:
            return True, rule.get('rule_name', '自動除外')
    return False, None


def match_category_rule(content, rules):
    """分類ルールにマッチするかチェック（優先度順にソート済み前提）"""
    for rule in rules:
        pattern = rule.get('content_pattern', '')
        if pattern and pattern in content:
            return {
                'cat_major': rule.get('category_major', ''),
                'cat_mid': rule.get('category_mid', ''),
                'cat_small': rule.get('category_small', ''),
                'cat_shop': rule.get('category_shop', ''),
                'cat_belonging': rule.get('category_belonging', ''),
                'cat_person': rule.get('category_person', ''),
                'cat_context': rule.get('category_context', ''),
                'rule_id': rule.get('id'),
                'rule_pattern': pattern
            }
    return None


def match_product_name_rule(ocr_name, shop_name, rules):
    """商品名ルールにマッチするかチェック（店舗別ルール優先）"""
    # 1. 店舗別ルールを優先
    for rule in rules:
        if rule.get('shop_name') == shop_name and rule.get('ocr_name') == ocr_name:
            return rule.get('product_name')

    # 2. 全店舗共通ルール
    for rule in rules:
        if rule.get('shop_name') is None and rule.get('ocr_name') == ocr_name:
            return rule.get('product_name')

    # 3. 部分一致（取得名を含むルール）
    for rule in rules:
        if rule.get('ocr_name') and rule.get('ocr_name') in ocr_name:
            return rule.get('product_name')

    return None


@app.route('/')
def index():
    """メイン画面：明細一覧（カードローンを除く）"""
    db = get_db()

    # フィルタパラメータ
    year_month = request.args.get('month', '')  # 例: 2025-12
    show_excluded = request.args.get('show_excluded', 'false') == 'true'

    # 自動除外ルールを取得
    rules_res = db.table("Kakeibo_Auto_Exclude_Rules").select("*").eq("is_active", True).execute()
    auto_exclude_rules = rules_res.data

    # 分類ルールを取得（優先度順）
    cat_rules_res = db.table("Kakeibo_Category_Rules").select("*").eq("is_active", True).order("priority", desc=True).order("use_count", desc=True).execute()
    category_rules = cat_rules_res.data

    # RawDataを取得
    query = db.table("Rawdata_BANK_transactions").select("*").order("date", desc=True)

    # 月フィルタ
    if year_month:
        start_date = f"{year_month}-01"
        year, month = map(int, year_month.split('-'))
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query = query.gte("date", start_date).lt("date", end_date)
    else:
        query = query.limit(500)

    res = query.execute()
    transactions = res.data
    ids = [t['id'] for t in transactions]

    # Manualデータを取得
    manual_map = {}
    if ids:
        m_res = db.table("Kakeibo_Manual_Edits").select("*").in_("transaction_id", ids).execute()
        manual_map = {m['transaction_id']: m for m in m_res.data}

    # カードローン取引のIDを取得（自動判定用）
    card_loan_ids = set()
    if ids:
        loan_res = db.table("view_kakeibo_loan_entries").select("transaction_id, loan_type").in_("transaction_id", ids).execute()
        card_loan_ids = {e['transaction_id'] for e in loan_res.data if e.get('loan_type') == 'card_loan'}

    # 表示用データ構築
    display_data = []
    for t in transactions:
        m = manual_map.get(t['id'], {})
        is_excluded = m.get("is_excluded", False)
        view_target = m.get("view_target")  # 'loan', 'list', or None

        # 自動除外ルールをチェック（手動で除外設定されていない場合のみ）
        auto_excluded, auto_rule = check_auto_exclude(
            t.get('content', ''),
            t.get('institution', ''),
            auto_exclude_rules
        )
        if auto_excluded and not m:
            is_excluded = True

        # 除外データを表示しない場合はスキップ
        if is_excluded and not show_excluded:
            continue

        # カードローン判定：view_target='loan' または (view_target=NoneかつカードローンID)
        is_card_loan = t['id'] in card_loan_ids
        if view_target == 'loan' or (view_target is None and is_card_loan):
            # ローン管理に表示するのでスキップ（view_target='list'で強制表示の場合は除く）
            if view_target != 'list':
                continue

        # 分類ルールから自動提案（手動設定がない場合のみ）
        has_manual_category = m.get("category_major") or m.get("category_mid")
        suggested = None
        if not has_manual_category:
            suggested = match_category_rule(t.get('content', ''), category_rules)

        display_data.append({
            **t,
            "cat_major": m.get("category_major", "") or (suggested.get('cat_major', '') if suggested else ''),
            "cat_mid": m.get("category_mid", "") or (suggested.get('cat_mid', '') if suggested else ''),
            "cat_small": m.get("category_small", "") or (suggested.get('cat_small', '') if suggested else ''),
            "cat_shop": m.get("category_shop", "") or (suggested.get('cat_shop', '') if suggested else ''),
            "cat_belonging": m.get("category_belonging", "") or (suggested.get('cat_belonging', '') if suggested else ''),
            "cat_person": m.get("category_person", "") or (suggested.get('cat_person', '') if suggested else ''),
            "cat_context": m.get("category_context", "") or (suggested.get('cat_context', '') if suggested else ''),
            "is_excluded": is_excluded,
            "is_auto_excluded": auto_excluded and not m,
            "auto_exclude_rule": auto_rule,
            "is_suggested": suggested is not None and not has_manual_category,
            "suggested_rule": suggested.get('rule_pattern') if suggested else None,
            "note": m.get("note", ""),
            "view_target": view_target,
            "is_card_loan": is_card_loan
        })

    # 集計
    total_expense = sum(t['amount'] for t in display_data if t['amount'] < 0 and not t['is_excluded'])
    total_income = sum(t['amount'] for t in display_data if t['amount'] > 0 and not t['is_excluded'])

    return render_template(
        'index.html',
        transactions=display_data,
        total_expense=total_expense,
        total_income=total_income,
        current_month=year_month,
        show_excluded=show_excluded
    )


@app.route('/api/update', methods=['POST'])
def update_transaction():
    """修正API：分類や除外フラグを更新"""
    data = request.json
    tx_id = data.get('id')

    if not tx_id:
        return jsonify({"status": "error", "message": "ID is required"}), 400

    payload = {
        "transaction_id": tx_id,
        "category_major": data.get('cat_major') or None,
        "category_mid": data.get('cat_mid') or None,
        "category_small": data.get('cat_small') or None,
        "category_shop": data.get('cat_shop') or None,
        "category_belonging": data.get('cat_belonging') or None,
        "category_person": data.get('cat_person') or None,
        "category_context": data.get('cat_context') or None,
        "is_excluded": data.get('is_excluded', False),
        "note": data.get('note') or None
    }

    db = get_db()
    # Upsert実行
    res = db.table("Kakeibo_Manual_Edits").upsert(payload, on_conflict="transaction_id").execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success"})


@app.route('/api/bulk_exclude', methods=['POST'])
def bulk_exclude():
    """一括除外API"""
    data = request.json
    ids = data.get('ids', [])

    if not ids:
        return jsonify({"status": "error", "message": "No IDs provided"}), 400

    db = get_db()
    for tx_id in ids:
        payload = {
            "transaction_id": tx_id,
            "is_excluded": True
        }
        db.table("Kakeibo_Manual_Edits").upsert(payload, on_conflict="transaction_id").execute()

    return jsonify({"status": "success", "count": len(ids)})


@app.route('/loans')
def loans_page():
    """ローン管理画面（カードローンのみ、取引先別にグループ化）"""
    db = get_db()

    # カードローン口座のみ取得
    accounts_res = db.table("Kakeibo_Loan_Accounts").select("*").eq("is_active", True).eq("loan_type", "card_loan").execute()
    accounts = accounts_res.data

    # カードローン残高
    card_balances_res = db.table("view_kakeibo_card_loan_balances").select("*").execute()
    card_balances = {b['loan_id']: b for b in card_balances_res.data}

    # カードローン取引明細を取得
    entries_res = db.table("view_kakeibo_loan_entries").select("*").eq("loan_type", "card_loan").order("date", desc=True).execute()
    all_entries = entries_res.data

    # Manual Editsからview_targetを取得
    entry_ids = [e['transaction_id'] for e in all_entries if e.get('transaction_id')]
    manual_map = {}
    if entry_ids:
        m_res = db.table("Kakeibo_Manual_Edits").select("transaction_id, view_target").in_("transaction_id", entry_ids).execute()
        manual_map = {m['transaction_id']: m.get('view_target') for m in m_res.data}

    # view_target='list'のものは除外（明細一覧に移動されたもの）
    entries = []
    for e in all_entries:
        vt = manual_map.get(e.get('transaction_id'))
        if vt == 'list':
            continue
        e['view_target'] = vt
        entries.append(e)

    # 取引先（loan_id）ごとにグループ化
    entries_by_loan = {}
    for e in entries:
        lid = e.get('loan_id', 'unknown')
        if lid not in entries_by_loan:
            entries_by_loan[lid] = []
        entries_by_loan[lid].append(e)

    # 口座ごとにデータを統合
    loan_data = []
    for acc in accounts:
        loan_id = acc['loan_id']
        balance_info = card_balances.get(loan_id, {})
        current_balance = balance_info.get('current_balance', 0)
        loan_entries = entries_by_loan.get(loan_id, [])

        loan_data.append({
            **acc,
            'current_balance': current_balance,
            'balance_date': balance_info.get('latest_date'),
            'entries': loan_entries,
            'entry_count': len(loan_entries)
        })

    return render_template(
        'loans.html',
        loans=loan_data
    )


@app.route('/reconcile')
def reconcile_page():
    """消込（突き合わせ）専用画面"""
    db = get_db()

    # 除外されていないデータのみ取得
    res = db.table("Rawdata_BANK_transactions").select("*").order("date", desc=True).limit(1000).execute()
    raw_data = res.data

    ids = [r['id'] for r in raw_data]
    excluded_ids = set()
    if ids:
        try:
            manual_res = db.table("Kakeibo_Manual_Edits").select("transaction_id, is_excluded").in_("transaction_id", ids).execute()
            excluded_ids = {m['transaction_id'] for m in manual_res.data if m.get('is_excluded')}
        except Exception as e:
            print(f"Manual Edits取得エラー: {e}")

    # 除外済みでないものだけリスト化
    candidates = [r for r in raw_data if r['id'] not in excluded_ids]

    return render_template('reconcile.html', transactions=candidates)


@app.route('/api/reconcile_execute', methods=['POST'])
def reconcile_execute():
    """
    消込実行API
    - 振替パターン（合計=0）: 左右両方を除外
    - 同額パターン（左=右）: 左だけ除外（右は残す）
    """
    data = request.json
    remove_ids = data.get('remove_ids', [])  # 左側
    keep_ids = data.get('keep_ids', [])      # 右側
    mode = data.get('mode', 'transfer')      # 'transfer' or 'same_amount'

    if not remove_ids:
        return jsonify({"status": "error", "message": "左側が選択されていません"}), 400

    db = get_db()

    updates = []

    # 左側は常に除外
    for tx_id in remove_ids:
        updates.append({
            "transaction_id": tx_id,
            "is_excluded": True,
            "note": "消込完了（左側）"
        })

    # 振替モードの場合のみ右側も除外
    if mode == 'transfer':
        for tx_id in keep_ids:
            updates.append({
                "transaction_id": tx_id,
                "is_excluded": True,
                "note": "消込完了（振替・右側）"
            })

    res = db.table("Kakeibo_Manual_Edits").upsert(updates, on_conflict="transaction_id").execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success", "count": len(updates)})


@app.route('/api/create_transaction', methods=['POST'])
def create_transaction():
    """
    新規明細作成API（右パネルに追加用）
    """
    data = request.json
    content = data.get('content', '新規明細')
    amount = data.get('amount', 0)
    date = data.get('date')
    institution = data.get('institution', '手動入力')

    if not date:
        return jsonify({"status": "error", "message": "日付が必要です"}), 400

    db = get_db()

    new_id = f"MANUAL-{uuid.uuid4()}"

    new_record = {
        "id": new_id,
        "date": date,
        "content": content,
        "amount": amount,
        "institution": institution,
        "category_major": "未分類",
        "memo": "消込用に手動作成",
        "is_target": True,
        "is_transfer": False
    }

    res = db.table("Rawdata_BANK_transactions").insert(new_record).execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success", "new_id": new_id, "record": new_record})


@app.route('/api/toggle_view_target', methods=['POST'])
def toggle_view_target():
    """
    明細一覧 ⇔ ローン管理の表示先切り替えAPI
    view_target: 'loan' | 'list' | null(自動判定に戻す)
    """
    data = request.json
    tx_id = data.get('id')
    new_target = data.get('view_target')  # 'loan', 'list', or None

    if not tx_id:
        return jsonify({"status": "error", "message": "ID is required"}), 400

    db = get_db()
    payload = {
        "transaction_id": tx_id,
        "view_target": new_target
    }
    res = db.table("Kakeibo_Manual_Edits").upsert(payload, on_conflict="transaction_id").execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success", "view_target": new_target})


@app.route('/api/export')
def export_data():
    """明細一覧をCSV/Excelでエクスポート"""
    db = get_db()

    # パラメータ
    year_month = request.args.get('month', '')
    show_excluded = request.args.get('show_excluded', 'false') == 'true'
    fmt = request.args.get('format', 'csv')  # csv or excel

    # 自動除外ルールを取得
    rules_res = db.table("Kakeibo_Auto_Exclude_Rules").select("*").eq("is_active", True).execute()
    auto_exclude_rules = rules_res.data

    # 分類ルールを取得
    cat_rules_res = db.table("Kakeibo_Category_Rules").select("*").eq("is_active", True).order("priority", desc=True).execute()
    category_rules = cat_rules_res.data

    # RawDataを取得
    query = db.table("Rawdata_BANK_transactions").select("*").order("date", desc=True)

    if year_month:
        start_date = f"{year_month}-01"
        year, month = map(int, year_month.split('-'))
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query = query.gte("date", start_date).lt("date", end_date)
    else:
        query = query.limit(2000)

    res = query.execute()
    transactions = res.data
    ids = [t['id'] for t in transactions]

    # Manualデータを取得
    manual_map = {}
    if ids:
        m_res = db.table("Kakeibo_Manual_Edits").select("*").in_("transaction_id", ids).execute()
        manual_map = {m['transaction_id']: m for m in m_res.data}

    # カードローン取引のIDを取得
    card_loan_ids = set()
    if ids:
        loan_res = db.table("view_kakeibo_loan_entries").select("transaction_id, loan_type").in_("transaction_id", ids).execute()
        card_loan_ids = {e['transaction_id'] for e in loan_res.data if e.get('loan_type') == 'card_loan'}

    # エクスポート用データ構築
    export_data = []
    for t in transactions:
        m = manual_map.get(t['id'], {})
        is_excluded = m.get("is_excluded", False)
        view_target = m.get("view_target")

        # 自動除外チェック
        auto_excluded, _ = check_auto_exclude(t.get('content', ''), t.get('institution', ''), auto_exclude_rules)
        if auto_excluded and not m:
            is_excluded = True

        if is_excluded and not show_excluded:
            continue

        # カードローンはスキップ
        is_card_loan = t['id'] in card_loan_ids
        if view_target == 'loan' or (view_target is None and is_card_loan):
            if view_target != 'list':
                continue

        # 分類ルールから提案
        has_manual = m.get("category_major") or m.get("category_mid")
        suggested = None
        if not has_manual:
            suggested = match_category_rule(t.get('content', ''), category_rules)

        export_data.append({
            'date': t.get('date', ''),
            'content': t.get('content', ''),
            'institution': t.get('institution', ''),
            'amount': t.get('amount', 0),
            'cat_major': m.get("category_major", "") or (suggested.get('cat_major', '') if suggested else ''),
            'cat_mid': m.get("category_mid", "") or (suggested.get('cat_mid', '') if suggested else ''),
            'cat_small': m.get("category_small", "") or (suggested.get('cat_small', '') if suggested else ''),
            'cat_shop': m.get("category_shop", "") or (suggested.get('cat_shop', '') if suggested else ''),
            'cat_belonging': m.get("category_belonging", "") or (suggested.get('cat_belonging', '') if suggested else ''),
            'cat_person': m.get("category_person", "") or (suggested.get('cat_person', '') if suggested else ''),
            'cat_context': m.get("category_context", "") or (suggested.get('cat_context', '') if suggested else ''),
            'note': m.get("note", ""),
            'is_excluded': '除外' if is_excluded else '',
        })

    # ヘッダー
    headers = ['日付', '内容', '金融機関', '金額', '大分類', '中分類', '小分類', '店', '所属', '人', '文脈', 'メモ', '除外']

    if fmt == 'excel' and HAS_OPENPYXL:
        return export_excel(export_data, headers, year_month)
    else:
        return export_csv(export_data, headers, year_month)


def export_csv(data, headers, year_month):
    """CSV出力"""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)

    # ヘッダー
    writer.writerow(headers)

    # データ
    for row in data:
        writer.writerow([
            row['date'],
            row['content'],
            row['institution'],
            row['amount'],
            row['cat_major'],
            row['cat_mid'],
            row['cat_small'],
            row['cat_shop'],
            row['cat_belonging'],
            row['cat_person'],
            row['cat_context'],
            row['note'],
            row['is_excluded'],
        ])

    output.seek(0)
    filename = f"kakeibo_{year_month or 'all'}_{datetime.now().strftime('%Y%m%d')}.csv"

    return Response(
        output.getvalue().encode('utf-8-sig'),  # BOM付きUTF-8（Excel対応）
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


def export_excel(data, headers, year_month):
    """Excel出力"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明細一覧"

    # ヘッダースタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # ヘッダー行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # データ行
    for row_idx, row in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=row['date'])
        ws.cell(row=row_idx, column=2, value=row['content'])
        ws.cell(row=row_idx, column=3, value=row['institution'])
        ws.cell(row=row_idx, column=4, value=row['amount'])
        ws.cell(row=row_idx, column=5, value=row['cat_major'])
        ws.cell(row=row_idx, column=6, value=row['cat_mid'])
        ws.cell(row=row_idx, column=7, value=row['cat_small'])
        ws.cell(row=row_idx, column=8, value=row['cat_shop'])
        ws.cell(row=row_idx, column=9, value=row['cat_belonging'])
        ws.cell(row=row_idx, column=10, value=row['cat_person'])
        ws.cell(row=row_idx, column=11, value=row['cat_context'])
        ws.cell(row=row_idx, column=12, value=row['note'])
        ws.cell(row=row_idx, column=13, value=row['is_excluded'])

    # 列幅調整
    col_widths = [12, 30, 15, 12, 10, 10, 10, 15, 10, 10, 12, 20, 6]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # フィルター設定
    ws.auto_filter.ref = ws.dimensions

    # 出力
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"kakeibo_{year_month or 'all'}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@app.route('/api/register_rule', methods=['POST'])
def register_rule():
    """
    分類ルール登録API
    現在の分類設定をルールとして保存
    """
    data = request.json
    pattern = data.get('pattern')

    if not pattern:
        return jsonify({"status": "error", "message": "パターンが必要です"}), 400

    payload = {
        "content_pattern": pattern,
        "category_major": data.get('cat_major') or None,
        "category_mid": data.get('cat_mid') or None,
        "category_small": data.get('cat_small') or None,
        "category_shop": data.get('cat_shop') or None,
        "category_belonging": data.get('cat_belonging') or None,
        "category_person": data.get('cat_person') or None,
        "category_context": data.get('cat_context') or None,
    }

    db = get_db()
    # Upsert（同じパターンがあれば更新）
    res = db.table("Kakeibo_Category_Rules").upsert(payload, on_conflict="content_pattern").execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success", "pattern": pattern})


@app.route('/api/increment_rule_usage', methods=['POST'])
def increment_rule_usage():
    """ルール使用回数をインクリメント（学習用）"""
    data = request.json
    pattern = data.get('pattern')

    if not pattern:
        return jsonify({"status": "error", "message": "パターンが必要です"}), 400

    db = get_db()
    # use_countをインクリメント
    db.rpc('increment_category_rule_usage', {'p_pattern': pattern}).execute()

    return jsonify({"status": "success"})


@app.route('/api/merge_execute', methods=['POST'])
def merge_execute():
    """
    合算（Merge）API
    指定されたID群を除外(is_excluded=True)し、
    その合計金額を持つ「新しい明細行」をRawdataに作成する。
    """
    data = request.json
    target_ids = data.get('ids', [])

    if len(target_ids) < 2:
        return jsonify({"status": "error", "message": "合算するには2つ以上の明細を選択してください"}), 400

    db = get_db()

    # 1. 選択された明細の情報を取得
    res = db.table("Rawdata_BANK_transactions").select("*").in_("id", target_ids).execute()
    rows = res.data

    if not rows:
        return jsonify({"status": "error", "message": "データが見つかりません"}), 404

    # 2. 新しい「合算行」のデータを作成
    total_amount = sum(r['amount'] for r in rows)
    base_date = min(r['date'] for r in rows)
    combined_content = " + ".join([r['content'] for r in rows])
    if len(combined_content) > 50:
        combined_content = combined_content[:50] + "..."

    institutions = list(set(r['institution'] for r in rows if r.get('institution')))
    new_institution = institutions[0] if len(institutions) == 1 else "合算データ"

    new_id = f"MERGE-{uuid.uuid4()}"

    new_record = {
        "id": new_id,
        "date": base_date,
        "content": f"【合算】{combined_content}",
        "amount": total_amount,
        "institution": new_institution,
        "category_major": "未分類",
        "memo": "UI上で手動合算",
        "is_target": True,
        "is_transfer": False
    }

    # 3. 新しい行を追加
    insert_res = db.table("Rawdata_BANK_transactions").insert(new_record).execute()
    if getattr(insert_res, 'error', None):
        return jsonify({"status": "error", "message": str(insert_res.error)}), 500

    # 4. 古い行を除外設定
    updates = []
    for tid in target_ids:
        updates.append({
            "transaction_id": tid,
            "is_excluded": True,
            "note": f"合算ID:{new_id} へ統合"
        })

    upsert_res = db.table("Kakeibo_Manual_Edits").upsert(updates, on_conflict="transaction_id").execute()

    if getattr(upsert_res, 'error', None):
        return jsonify({"status": "error", "message": "除外処理に失敗しました"}), 500

    return jsonify({"status": "success", "new_id": new_id})


# ============================================================
# レシート検証画面
# ============================================================

@app.route('/receipts')
def receipts_page():
    """レシート一覧画面"""
    db = get_db()

    # フィルタパラメータ
    status_filter = request.args.get('status', 'all')  # all, success, failed, unverified

    # 処理ログ取得（レシート単位）
    query = db.table("99_lg_image_proc_log").select("*").order("processed_at", desc=True).limit(100)

    if status_filter == 'success':
        query = query.eq("status", "success")
    elif status_filter == 'failed':
        query = query.eq("status", "failed")

    logs = query.execute()

    # 紐付け済み件数を取得
    linked_receipts = set()
    if logs.data:
        receipt_ids = [log.get('receipt_id') for log in logs.data if log.get('receipt_id')]
        if receipt_ids:
            try:
                links_res = db.table("Kakeibo_Receipt_Links").select("receipt_id").in_("receipt_id", receipt_ids).execute()
                linked_receipts = {link['receipt_id'] for link in links_res.data}
            except Exception:
                # テーブルが存在しない場合は空
                pass

    # 表示用データ構築
    display_data = []
    for log in logs.data:
        status_icon = {"success": "✓", "failed": "✗"}.get(log.get("status"), "?")
        is_linked = log.get('receipt_id') in linked_receipts

        display_data.append({
            **log,
            "status_icon": status_icon,
            "is_linked": is_linked,
            "date_short": log.get("processed_at", "")[:10] if log.get("processed_at") else ""
        })

    return render_template(
        'receipts.html',
        receipts=display_data,
        status_filter=status_filter
    )


@app.route('/receipts/<receipt_id>')
def receipt_detail(receipt_id):
    """レシート詳細画面"""
    db = get_db()

    # 日付フィルタパラメータ
    filter_date = request.args.get('filter_date', '')

    # 処理ログを取得
    log_res = db.table("99_lg_image_proc_log").select("*").eq("receipt_id", receipt_id).execute()
    if not log_res.data:
        return "レシートが見つかりません", 404

    log = log_res.data[0]

    # レシート情報を取得
    receipt = None
    items = []
    if log.get("status") == "success" and log.get("receipt_id"):
        receipt_res = db.table("Rawdata_RECEIPT_shops").select("*").eq("id", receipt_id).execute()
        if receipt_res.data:
            receipt = receipt_res.data[0]

            # 明細を取得（ITEMのみ）
            items_res = db.table("Rawdata_RECEIPT_items").select("*").eq("receipt_id", receipt_id).order("line_number").execute()
            items = items_res.data

    # 画像をBase64エンコード
    image_base64 = None
    if log.get("drive_file_id"):
        image_base64 = get_receipt_image_base64(log["drive_file_id"])

    # 紐付け済み銀行取引を取得
    linked_transaction = None
    try:
        link_res = db.table("Kakeibo_Receipt_Links").select("*").eq("receipt_id", receipt_id).execute()
        if link_res.data:
            link_data = link_res.data[0]
            # 紐付けられた銀行取引を取得
            tx_res = db.table("Rawdata_BANK_transactions").select("*").eq("id", link_data['transaction_id']).execute()
            if tx_res.data:
                linked_transaction = {
                    **link_data,
                    "Rawdata_BANK_transactions": tx_res.data[0]
                }
    except Exception:
        # テーブルが存在しない場合
        pass

    # 紐付け候補となる取引を取得（日付同日優先）
    candidates = []
    receipt_date_str = ""
    from datetime import timedelta

    # 既にレシート紐付け済みの取引IDを取得（これらは候補から除外）
    linked_tx_ids = set()
    try:
        linked_res = db.table("Kakeibo_Receipt_Links").select("transaction_id").execute()
        linked_tx_ids = {l['transaction_id'] for l in linked_res.data if l.get('transaction_id')}
    except Exception:
        pass

    try:
        # 日付フィルタが指定されている場合はその日付のみ取得
        if filter_date:
            cand_res = db.table("Rawdata_BANK_transactions").select("*") \
                .eq("date", filter_date).execute()
            candidates = [c for c in cand_res.data if c['id'] not in linked_tx_ids]
            # レシート金額に近い順でソート
            if receipt and receipt.get("total_amount_check"):
                receipt_total = abs(receipt.get("total_amount_check", 0))
                candidates = sorted(candidates, key=lambda x: abs(abs(x.get('amount', 0)) - receipt_total))
        elif receipt and receipt.get("transaction_date"):
            tx_date = receipt["transaction_date"]
            receipt_date_str = tx_date[:10] if len(str(tx_date)) > 10 else str(tx_date)
            # レシート日付の前後60日間の取引を候補として取得
            base_date = datetime.strptime(receipt_date_str, "%Y-%m-%d")
            start_date = (base_date - timedelta(days=60)).strftime("%Y-%m-%d")
            end_date = (base_date + timedelta(days=60)).strftime("%Y-%m-%d")

            cand_res = db.table("Rawdata_BANK_transactions").select("*") \
                .gte("date", start_date).lte("date", end_date) \
                .order("date", desc=True).limit(500).execute()
            candidates = [c for c in cand_res.data if c['id'] not in linked_tx_ids]

            # 日付同日を優先し、同日内は金額近い順でソート
            def sort_key(x):
                x_date = str(x.get('date', ''))[:10]
                is_same_day = 0 if x_date == receipt_date_str else 1
                receipt_total = abs(receipt.get("total_amount_check", 0)) if receipt else 0
                amount_diff = abs(abs(x.get('amount', 0)) - receipt_total)
                return (is_same_day, amount_diff)
            candidates = sorted(candidates, key=sort_key)[:100]
        else:
            # 日付がない場合は最新100件
            cand_res = db.table("Rawdata_BANK_transactions").select("*") \
                .order("date", desc=True).limit(100).execute()
            candidates = [c for c in cand_res.data if c['id'] not in linked_tx_ids]

    except Exception as e:
        print(f"候補取得エラー: {e}")
        # フォールバック
        try:
            cand_res = db.table("Rawdata_BANK_transactions").select("*") \
                .order("date", desc=True).limit(100).execute()
            candidates = [c for c in cand_res.data if c['id'] not in linked_tx_ids]
        except:
            candidates = []

    # 商品名ルールを取得
    product_name_rules = []
    try:
        pn_rules_res = db.table("Kakeibo_Product_Name_Rules").select("*").eq("is_active", True).order("priority", desc=True).order("use_count", desc=True).execute()
        product_name_rules = pn_rules_res.data
    except Exception:
        pass

    shop_name = receipt.get("shop_name", "") if receipt else ""

    # 明細データの整形
    display_items = []
    for item in items:
        # 税込価を計算
        tax_included = item.get("std_amount") or item.get("tax_included_amount") or 0
        tax_amount = item.get("tax_amount") or 0
        base_price = tax_included - tax_amount if tax_included and tax_amount else item.get("base_price") or 0

        # 商品名を提案（official_nameが空の場合）
        ocr_name = item.get("product_name", "")
        current_product_name = item.get("official_name") or ""
        suggested_product_name = None
        if not current_product_name:
            suggested_product_name = match_product_name_rule(ocr_name, shop_name, product_name_rules)

        display_items.append({
            **item,
            "tax_included": tax_included,
            "base_price": base_price,
            "tax_rate": item.get("tax_rate") or 10,
            "ocr_name": ocr_name,  # 取得名（OCR）
            "product_name_display": current_product_name or suggested_product_name or "",  # 表示用商品名
            "is_suggested": suggested_product_name is not None and not current_product_name
        })

    # 合計金額計算（明細から合算）
    calc_total = sum(i.get("tax_included") or 0 for i in display_items)
    calc_tax = sum(i.get("tax_amount") or 0 for i in display_items)
    calc_tax_8 = sum(i.get("tax_amount") or 0 for i in display_items if i.get("tax_rate") == 8)
    calc_tax_10 = sum(i.get("tax_amount") or 0 for i in display_items if i.get("tax_rate") == 10)

    # レシートから読み取った値（OCR値）
    ocr_total = receipt.get("total_amount_check") if receipt else None
    ocr_subtotal = receipt.get("subtotal_amount") if receipt else None
    ocr_tax_8 = receipt.get("tax_8_amount") if receipt else None
    ocr_tax_10 = receipt.get("tax_10_amount") if receipt else None

    return render_template(
        'receipt_detail.html',
        log=log,
        receipt=receipt,
        items=display_items,
        image_base64=image_base64,
        linked_transaction=linked_transaction,
        candidates=candidates,
        filter_date=filter_date,
        # 計算値
        calc_total=calc_total,
        calc_tax=calc_tax,
        calc_tax_8=calc_tax_8,
        calc_tax_10=calc_tax_10,
        # OCR読み取り値
        ocr_total=ocr_total,
        ocr_subtotal=ocr_subtotal,
        ocr_tax_8=ocr_tax_8,
        ocr_tax_10=ocr_tax_10
    )


@app.route('/api/receipts/import', methods=['POST'])
def import_receipts():
    """Google Driveからレシート画像を取り込んで処理"""
    import json
    import tempfile
    from pathlib import Path

    # 設定読み込み
    try:
        from shared.kakeibo.config import (
            INBOX_EASY_FOLDER_ID, INBOX_HARD_FOLDER_ID,
            ARCHIVE_FOLDER_ID, ERROR_FOLDER_ID,
            GEMINI_MODEL_EASY, GEMINI_MODEL_HARD, GEMINI_PROMPT,
            GEMINI_TEMPERATURE
        )
        from shared.common.connectors.google_drive import GoogleDriveConnector
        from shared.ai.llm_client.llm_client import LLMClient
        from shared.kakeibo.transaction_processor import TransactionProcessor
    except ImportError as e:
        return jsonify({"status": "error", "message": f"必要なモジュールが見つかりません: {e}"}), 500

    # 処理結果カウンター
    processed = 0
    success = 0
    failed = 0
    errors = []

    try:
        drive = GoogleDriveConnector()
        llm = LLMClient()
        processor = TransactionProcessor()

        # 処理対象のフォルダ
        folders = [
            {"id": INBOX_EASY_FOLDER_ID, "model": GEMINI_MODEL_EASY, "name": "INBOX_EASY"},
            {"id": INBOX_HARD_FOLDER_ID, "model": GEMINI_MODEL_HARD, "name": "INBOX_HARD"},
        ]

        for folder in folders:
            if not folder["id"]:
                continue

            # フォルダ内のファイルを取得（画像のみ）
            try:
                files = drive.list_files_in_folder(folder["id"])
                # 画像ファイルのみフィルタ
                image_files = [
                    f for f in files
                    if f.get('mimeType', '').startswith('image/')
                ]
            except Exception as e:
                errors.append(f"フォルダ取得エラー ({folder['name']}): {e}")
                continue

            for file_info in image_files:
                file_id = file_info['id']
                file_name = file_info['name']
                processed += 1

                try:
                    # 一時ディレクトリにダウンロード
                    with tempfile.TemporaryDirectory() as temp_dir:
                        local_path = drive.download_file(file_id, file_name, temp_dir)
                        if not local_path:
                            raise Exception("ダウンロード失敗")

                        # 画像をBase64エンコード
                        with open(local_path, 'rb') as f:
                            image_bytes = f.read()
                        image_base64 = base64.b64encode(image_bytes).decode('utf-8')

                        # Gemini OCR実行
                        ocr_response = llm.generate_with_images(
                            prompt=GEMINI_PROMPT,
                            image_data=image_base64,
                            model=folder["model"],
                            temperature=GEMINI_TEMPERATURE
                        )

                        # JSON部分を抽出
                        json_start = ocr_response.find('{')
                        json_end = ocr_response.rfind('}') + 1
                        if json_start == -1 or json_end == 0:
                            raise Exception("JSONが見つかりません")

                        ocr_result = json.loads(ocr_response[json_start:json_end])

                        # OCR結果に必要なフィールドがあるか確認
                        if "error" in ocr_result:
                            raise Exception(f"OCRエラー: {ocr_result.get('message', ocr_result['error'])}")

                        # transaction_date を transaction_info から取得
                        if "transaction_info" in ocr_result and "date" in ocr_result["transaction_info"]:
                            ocr_result["transaction_date"] = ocr_result["transaction_info"]["date"]

                        # shop_name を shop_info から取得
                        if "shop_info" in ocr_result and "name" in ocr_result["shop_info"]:
                            ocr_result["shop_name"] = ocr_result["shop_info"]["name"]

                        # items の変換（Gemini出力 → TransactionProcessor入力形式）
                        if "items" in ocr_result:
                            for item in ocr_result["items"]:
                                if "line_type" not in item:
                                    item["line_type"] = "ITEM"
                                if "line_text" not in item and "product_name" in item:
                                    item["line_text"] = item["product_name"]

                        # amounts から合計金額を取得
                        if "amounts" in ocr_result:
                            amounts = ocr_result["amounts"]
                            ocr_result["subtotal_amount"] = amounts.get("subtotal")
                            ocr_result["tax_8_amount"] = amounts.get("tax_8_amount")
                            ocr_result["tax_10_amount"] = amounts.get("tax_10_amount")
                            ocr_result["total_amount_check"] = amounts.get("total")

                        # TransactionProcessor で処理
                        result = processor.process(
                            ocr_result=ocr_result,
                            file_name=file_name,
                            drive_file_id=file_id,
                            model_name=folder["model"],
                            source_folder=folder["name"]
                        )

                        if "error" in result:
                            raise Exception(result.get("message", result["error"]))

                        # 成功時はアーカイブに移動
                        if ARCHIVE_FOLDER_ID:
                            try:
                                drive.move_file(file_id, ARCHIVE_FOLDER_ID)
                            except Exception:
                                pass  # 移動失敗は無視

                        success += 1

                except Exception as e:
                    failed += 1
                    error_msg = f"{file_name}: {str(e)}"
                    errors.append(error_msg)

                    # エラー時はエラーフォルダに移動
                    if ERROR_FOLDER_ID:
                        try:
                            drive.move_file(file_id, ERROR_FOLDER_ID)
                        except Exception:
                            pass

        return jsonify({
            "status": "success",
            "processed": processed,
            "success": success,
            "failed": failed,
            "errors": errors[:10]  # 最初の10件のみ
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/receipt/image/<drive_file_id>')
def get_receipt_image(drive_file_id):
    """レシート画像を取得（API）"""
    image_base64 = get_receipt_image_base64(drive_file_id)
    if image_base64:
        return jsonify({"status": "success", "image": image_base64})
    return jsonify({"status": "error", "message": "画像を取得できませんでした"}), 404


@app.route('/api/receipt/link', methods=['POST'])
def link_receipt_to_transaction():
    """レシートと取引を紐付け（取引は除外される）"""
    data = request.json
    receipt_id = data.get('receipt_id')
    transaction_id = data.get('transaction_id')

    if not receipt_id or not transaction_id:
        return jsonify({"status": "error", "message": "receipt_id and transaction_id are required"}), 400

    db = get_db()

    # 既存の紐付けがあれば、その取引の除外を解除
    try:
        old_link = db.table("Kakeibo_Receipt_Links").select("transaction_id").eq("receipt_id", receipt_id).execute()
        if old_link.data:
            old_tx_id = old_link.data[0]['transaction_id']
            db.table("Kakeibo_Manual_Edits").upsert({
                "transaction_id": old_tx_id,
                "is_excluded": False,
                "has_receipt": False,
                "receipt_id": None,
                "note": "レシート紐付け解除"
            }, on_conflict="transaction_id").execute()
    except Exception:
        pass

    # 既存の紐付けを削除
    try:
        db.table("Kakeibo_Receipt_Links").delete().eq("receipt_id", receipt_id).execute()
    except Exception:
        pass

    # 新規紐付け
    payload = {
        "receipt_id": receipt_id,
        "transaction_id": transaction_id
    }
    try:
        res = db.table("Kakeibo_Receipt_Links").insert(payload).execute()
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

    # 取引を除外（レシート明細がデータになるため）
    db.table("Kakeibo_Manual_Edits").upsert({
        "transaction_id": transaction_id,
        "is_excluded": True,
        "has_receipt": True,
        "receipt_id": receipt_id,
        "note": "レシート紐付け（レシート明細に置換）"
    }, on_conflict="transaction_id").execute()

    # レシート明細を明細一覧に追加
    try:
        # レシート情報を取得
        receipt = db.table("Rawdata_RECEIPT_shops").select("shop_name, transaction_date").eq("id", receipt_id).execute()
        if receipt.data:
            shop_name = receipt.data[0].get("shop_name", "")
            tx_date = receipt.data[0].get("transaction_date")

            # 明細を取得
            items = db.table("Rawdata_RECEIPT_items").select("*").eq("receipt_id", receipt_id).eq("line_type", "ITEM").execute()

            for item in items.data:
                # 商品名: official_nameがあればそれ、なければproduct_name
                product_name = item.get("official_name") or item.get("product_name", "")
                amount = item.get("std_amount") or 0

                new_id = f"RECEIPT-{receipt_id}-{item['id']}"

                # 既存チェック（重複防止）
                existing = db.table("Rawdata_BANK_transactions").select("id").eq("id", new_id).execute()
                if existing.data:
                    continue

                new_record = {
                    "id": new_id,
                    "date": tx_date,
                    "content": product_name,
                    "amount": amount,
                    "institution": shop_name,
                    "category_major": "未分類",
                    "memo": f"レシート明細（{receipt_id}）",
                    "is_target": True,
                    "is_transfer": False
                }
                db.table("Rawdata_BANK_transactions").insert(new_record).execute()
    except Exception as e:
        print(f"レシート明細の追加に失敗: {e}")

    return jsonify({"status": "success"})


@app.route('/api/receipt/unlink', methods=['POST'])
def unlink_receipt():
    """レシートと取引の紐付けを解除（取引の除外も解除、追加した明細も削除）"""
    data = request.json
    receipt_id = data.get('receipt_id')

    if not receipt_id:
        return jsonify({"status": "error", "message": "receipt_id is required"}), 400

    db = get_db()

    # 紐付けられていた取引の除外を解除
    try:
        link_res = db.table("Kakeibo_Receipt_Links").select("transaction_id").eq("receipt_id", receipt_id).execute()
        if link_res.data:
            tx_id = link_res.data[0]['transaction_id']
            db.table("Kakeibo_Manual_Edits").upsert({
                "transaction_id": tx_id,
                "is_excluded": False,
                "has_receipt": False,
                "receipt_id": None,
                "note": "レシート紐付け解除"
            }, on_conflict="transaction_id").execute()
    except Exception:
        pass

    # 紐付けを削除
    try:
        db.table("Kakeibo_Receipt_Links").delete().eq("receipt_id", receipt_id).execute()
    except Exception:
        pass

    # レシート明細から追加された取引を削除
    try:
        db.table("Rawdata_BANK_transactions").delete().like("id", f"RECEIPT-{receipt_id}-%").execute()
    except Exception:
        pass

    return jsonify({"status": "success"})


@app.route('/api/receipt/update_item', methods=['POST'])
def update_receipt_item():
    """レシート明細を更新"""
    data = request.json
    item_id = data.get('item_id')

    if not item_id:
        return jsonify({"status": "error", "message": "item_id is required"}), 400

    db = get_db()

    # 更新データを構築
    update_data = {}
    fields = ['product_name', 'quantity', 'tax_rate', 'base_price', 'tax_amount', 'std_amount',
              'std_unit_price', 'minor_category', 'person', 'purpose', 'official_name']
    for field in fields:
        if field in data:
            update_data[field] = data[field]

    if not update_data:
        return jsonify({"status": "error", "message": "更新するデータがありません"}), 400

    res = db.table("Rawdata_RECEIPT_items").update(update_data).eq("id", item_id).execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success"})


@app.route('/api/receipt/verify', methods=['POST'])
def verify_receipt():
    """レシートを確認済みにする（商品名ルールも登録）"""
    data = request.json
    receipt_id = data.get('receipt_id')

    if not receipt_id:
        return jsonify({"status": "error", "message": "receipt_id is required"}), 400

    db = get_db()

    # レシート情報を取得（店舗名用）
    receipt = db.table("Rawdata_RECEIPT_shops").select("shop_name").eq("id", receipt_id).execute()
    shop_name = receipt.data[0].get("shop_name") if receipt.data else None

    # 明細を取得
    items = db.table("Rawdata_RECEIPT_items").select("*").eq("receipt_id", receipt_id).eq("line_type", "ITEM").execute()

    for item in items.data:
        ocr_name = item.get("product_name", "")
        official_name = item.get("official_name")

        # official_nameが空の場合、product_name（取得名）をコピー
        if not official_name:
            db.table("Rawdata_RECEIPT_items").update({
                "official_name": ocr_name
            }).eq("id", item["id"]).execute()
        else:
            # 取得名と商品名が異なる場合、ルールを登録
            if ocr_name and official_name != ocr_name:
                try:
                    # 同じルールが既に存在するかチェック
                    existing = db.table("Kakeibo_Product_Name_Rules").select("id, use_count").eq("ocr_name", ocr_name).eq("product_name", official_name).execute()
                    if existing.data:
                        # 使用回数を増やす
                        db.table("Kakeibo_Product_Name_Rules").update({
                            "use_count": (existing.data[0].get("use_count") or 0) + 1
                        }).eq("id", existing.data[0]["id"]).execute()
                    else:
                        # 新規ルールを登録
                        db.table("Kakeibo_Product_Name_Rules").insert({
                            "ocr_name": ocr_name,
                            "product_name": official_name,
                            "shop_name": shop_name
                        }).execute()
                except Exception:
                    pass  # ルール登録失敗は無視

    # レシートを確認済みに
    res = db.table("Rawdata_RECEIPT_shops").update({"is_verified": True}).eq("id", receipt_id).execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success"})


@app.route('/api/receipt/update_header', methods=['POST'])
def update_receipt_header():
    """レシートヘッダー（店舗情報）を更新"""
    data = request.json
    receipt_id = data.get('receipt_id')

    if not receipt_id:
        return jsonify({"status": "error", "message": "receipt_id is required"}), 400

    db = get_db()

    # 更新データを構築
    update_data = {}
    fields = ['shop_name', 'transaction_date', 'total_amount_check', 'subtotal_amount',
              'tax_8_amount', 'tax_10_amount', 'tax_8_subtotal', 'tax_10_subtotal']
    for field in fields:
        if field in data:
            update_data[field] = data[field]

    if not update_data:
        return jsonify({"status": "error", "message": "更新するデータがありません"}), 400

    res = db.table("Rawdata_RECEIPT_shops").update(update_data).eq("id", receipt_id).execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success"})


@app.route('/api/receipt/add_item', methods=['POST'])
def add_receipt_item():
    """レシート明細を追加"""
    data = request.json
    receipt_id = data.get('receipt_id')
    product_name = data.get('product_name')

    if not receipt_id or not product_name:
        return jsonify({"status": "error", "message": "receipt_id and product_name are required"}), 400

    db = get_db()

    # 最大行番号を取得
    items = db.table("Rawdata_RECEIPT_items").select("line_number").eq("receipt_id", receipt_id).order("line_number", desc=True).limit(1).execute()
    max_line = items.data[0]['line_number'] if items.data else 0

    # 新規明細を作成
    new_item = {
        "receipt_id": receipt_id,
        "line_number": max_line + 1,
        "line_type": "ITEM",
        "product_name": product_name,
        "official_name": data.get('official_name'),  # 商品名（手入力）
        "quantity": data.get('quantity', 1),
        "std_amount": data.get('std_amount', 0),
        "tax_amount": data.get('tax_amount', 0),
        "tax_rate": data.get('tax_rate', 10),
        "person": data.get('person'),
        "purpose": data.get('purpose')
    }

    res = db.table("Rawdata_RECEIPT_items").insert(new_item).execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success", "item": res.data[0] if res.data else None})


@app.route('/api/receipt/delete_item', methods=['POST'])
def delete_receipt_item():
    """レシート明細を削除"""
    data = request.json
    item_id = data.get('item_id')

    if not item_id:
        return jsonify({"status": "error", "message": "item_id is required"}), 400

    db = get_db()
    res = db.table("Rawdata_RECEIPT_items").delete().eq("id", item_id).execute()

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success"})


@app.route('/api/receipt/delete', methods=['POST'])
def delete_receipt():
    """レシート全体を削除（紐付けも解除）"""
    data = request.json
    receipt_id = data.get('receipt_id')

    if not receipt_id:
        return jsonify({"status": "error", "message": "receipt_id is required"}), 400

    db = get_db()

    # 紐付けを解除（取引の除外も解除）
    try:
        link_res = db.table("Kakeibo_Receipt_Links").select("transaction_id").eq("receipt_id", receipt_id).execute()
        if link_res.data:
            tx_id = link_res.data[0]['transaction_id']
            db.table("Kakeibo_Manual_Edits").upsert({
                "transaction_id": tx_id,
                "is_excluded": False,
                "has_receipt": False,
                "receipt_id": None,
                "note": "レシート削除により紐付け解除"
            }, on_conflict="transaction_id").execute()
        db.table("Kakeibo_Receipt_Links").delete().eq("receipt_id", receipt_id).execute()
    except Exception:
        pass

    # 明細を削除（CASCADE設定があれば自動削除されるが念のため）
    db.table("Rawdata_RECEIPT_items").delete().eq("receipt_id", receipt_id).execute()

    # レシート本体を削除
    res = db.table("Rawdata_RECEIPT_shops").delete().eq("id", receipt_id).execute()

    # 処理ログも更新
    try:
        db.table("99_lg_image_proc_log").update({"status": "deleted", "receipt_id": None}).eq("receipt_id", receipt_id).execute()
    except Exception:
        pass

    if getattr(res, 'error', None):
        return jsonify({"status": "error", "message": str(res.error)}), 500

    return jsonify({"status": "success"})


@app.route('/api/receipt/merge_item', methods=['POST'])
def merge_receipt_item():
    """割引行を上の行と合算"""
    data = request.json
    item_id = data.get('item_id')
    receipt_id = data.get('receipt_id')

    if not item_id or not receipt_id:
        return jsonify({"status": "error", "message": "item_id and receipt_id are required"}), 400

    db = get_db()

    # 対象の明細を取得
    item_res = db.table("Rawdata_RECEIPT_items").select("*").eq("id", item_id).execute()
    if not item_res.data:
        return jsonify({"status": "error", "message": "明細が見つかりません"}), 404

    current_item = item_res.data[0]
    current_line = current_item.get("line_number", 0)

    # 上の行を取得（line_numberが1つ小さい行）
    above_res = db.table("Rawdata_RECEIPT_items").select("*").eq("receipt_id", receipt_id).eq("line_type", "ITEM").lt("line_number", current_line).order("line_number", desc=True).limit(1).execute()

    if not above_res.data:
        return jsonify({"status": "error", "message": "上の行がありません"}), 400

    above_item = above_res.data[0]

    # 金額を合算（上の行に割引を加算）
    new_std_amount = (above_item.get("std_amount") or 0) + (current_item.get("std_amount") or 0)
    new_tax_amount = (above_item.get("tax_amount") or 0) + (current_item.get("tax_amount") or 0)

    # 上の行を更新
    db.table("Rawdata_RECEIPT_items").update({
        "std_amount": new_std_amount,
        "tax_amount": new_tax_amount
    }).eq("id", above_item["id"]).execute()

    # 割引行を削除
    db.table("Rawdata_RECEIPT_items").delete().eq("id", item_id).execute()

    return jsonify({"status": "success"})


@app.route('/api/receipt/register_product_rule', methods=['POST'])
def register_product_rule():
    """商品名ルールを登録"""
    data = request.json
    ocr_name = data.get('ocr_name')
    product_name = data.get('product_name')
    shop_name = data.get('shop_name')

    if not ocr_name or not product_name:
        return jsonify({"status": "error", "message": "ocr_name and product_name are required"}), 400

    db = get_db()

    # 同じルールが既に存在するかチェック
    existing = db.table("Kakeibo_Product_Name_Rules").select("id, use_count").eq("ocr_name", ocr_name).execute()

    if existing.data:
        # 既存ルールを更新
        db.table("Kakeibo_Product_Name_Rules").update({
            "product_name": product_name,
            "shop_name": shop_name,
            "use_count": (existing.data[0].get("use_count") or 0) + 1
        }).eq("id", existing.data[0]["id"]).execute()
    else:
        # 新規ルールを登録
        db.table("Kakeibo_Product_Name_Rules").insert({
            "ocr_name": ocr_name,
            "product_name": product_name,
            "shop_name": shop_name
        }).execute()

    return jsonify({"status": "success"})


@app.route('/api/receipt/mark_cash', methods=['POST'])
def mark_receipt_as_cash():
    """現金決済として登録（紐付けなしで明細一覧へ追加）"""
    data = request.json
    receipt_id = data.get('receipt_id')

    if not receipt_id:
        return jsonify({"status": "error", "message": "receipt_id is required"}), 400

    db = get_db()

    # レシート情報を取得
    receipt = db.table("Rawdata_RECEIPT_shops").select("shop_name, transaction_date").eq("id", receipt_id).execute()
    if not receipt.data:
        return jsonify({"status": "error", "message": "レシートが見つかりません"}), 404

    shop_name = receipt.data[0].get("shop_name", "")
    tx_date = receipt.data[0].get("transaction_date")

    # 明細を取得
    items = db.table("Rawdata_RECEIPT_items").select("*").eq("receipt_id", receipt_id).eq("line_type", "ITEM").execute()

    # 明細を明細一覧に追加
    for item in items.data:
        product_name = item.get("official_name") or item.get("product_name", "")
        amount = item.get("std_amount") or 0

        new_id = f"CASH-{receipt_id}-{item['id']}"

        # 既存チェック（重複防止）
        existing = db.table("Rawdata_BANK_transactions").select("id").eq("id", new_id).execute()
        if existing.data:
            continue

        new_record = {
            "id": new_id,
            "date": tx_date,
            "content": product_name,
            "amount": amount,
            "institution": "現金",
            "category_major": "未分類",
            "memo": f"現金決済レシート（{shop_name}）",
            "is_target": True,
            "is_transfer": False
        }
        db.table("Rawdata_BANK_transactions").insert(new_record).execute()

    # レシートを確認済みに
    db.table("Rawdata_RECEIPT_shops").update({"is_verified": True, "is_cash": True}).eq("id", receipt_id).execute()

    return jsonify({"status": "success"})


if __name__ == '__main__':
    app.run(debug=True, port=5000)
