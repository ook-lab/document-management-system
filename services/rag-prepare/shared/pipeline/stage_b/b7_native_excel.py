"""
B-7: Native Excel Processor (.xlsx専用)

openpyxl / pandas を使用して、ネイティブExcelファイルから構造化データを抽出。
シート・セル構造を100%精度で取得、背景色・フォント色も保持。
"""

from pathlib import Path
from typing import Dict, Any, List
from loguru import logger


class B7NativeExcelProcessor:
    """B-7: Native Excel Processor (.xlsx専用)"""

    def process(self, file_path: Path, log_file=None) -> Dict[str, Any]:
        """
        .xlsx ファイルから構造化データを抽出

        Args:
            file_path: .xlsxファイルパス
            log_file: 個別ログファイルパス（Noneなら共有ロガーのみ）

        Returns:
            {
                'is_structured': bool,
                'text_with_tags': str,           # 全シートのテキスト
                'structured_tables': [...],      # シートごとの表データ
                'sheets': [...],                 # シート情報
                'tags': {...},                   # メタ情報
                'media_elements': [...]          # 埋め込み画像
            }
        """
        _sink_id = None
        if log_file:
            _sink_id = logger.add(
                str(log_file),
                format="{time:HH:mm:ss} | {level:<5} | {message}",
                filter=lambda r: "[B-7]" in r["message"],
                level="DEBUG",
                encoding="utf-8",
            )
        try:
            return self._process_impl(file_path)
        finally:
            if _sink_id is not None:
                logger.remove(_sink_id)

    def _process_impl(self, file_path: Path) -> Dict[str, Any]:
        logger.info(f"[B-7] Native Excel処理開始: {file_path.name}")

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Color
        except ImportError:
            logger.error("[B-7] openpyxl がインストールされていません")
            return self._error_result("openpyxl not installed")

        try:
            wb = load_workbook(str(file_path), data_only=True)

            # 全シートを抽出
            sheets = self._extract_sheets(wb)

            # テキストを生成
            text_with_tags = self._build_text(sheets)

            # メタ情報
            tags = {
                'sheet_count': len(sheets),
                'has_merged_cells': any(s.get('has_merged', False) for s in sheets),
                'total_rows': sum(s.get('max_row', 0) for s in sheets),
                'total_cols': sum(s.get('max_col', 0) for s in sheets)
            }

            logger.info(f"[B-7] 抽出完了: シート={len(sheets)}")
            for sheet in sheets:
                logger.info(f"[B-7] sheet '{sheet.get('name', '')}': {sheet.get('max_row', 0)}行 x {sheet.get('max_col', 0)}列")
                for row_idx, row in enumerate(sheet.get('data', [])):
                    logger.info(f"[B-7] sheet '{sheet.get('name', '')}' 行{row_idx}: {row}")

            return {
                'is_structured': True,
                'text_with_tags': text_with_tags,
                'structured_tables': sheets,  # シートごとの表データ
                'sheets': sheets,
                'tags': tags,
                'media_elements': []  # TODO: 画像抽出
            }

        except Exception as e:
            logger.error(f"[B-7] 処理エラー: {e}", exc_info=True)
            return self._error_result(str(e))

    def _extract_sheets(self, wb) -> List[Dict[str, Any]]:
        """
        全シートを抽出

        Returns:
            [{
                'name': str,
                'max_row': int,
                'max_col': int,
                'data': [[...], [...], ...],
                'cell_properties': {...},
                'has_merged': bool
            }, ...]
        """
        sheets = []

        for sheet in wb.worksheets:
            # セルデータを抽出
            data = []
            cell_properties = {}
            has_merged = len(sheet.merged_cells.ranges) > 0

            for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                row_data = []
                for col_idx, cell in enumerate(row, start=1):
                    cell_value = cell.value if cell.value is not None else ""
                    row_data.append(str(cell_value))

                    # セルのプロパティを保存
                    cell_addr = f"{sheet.title}!{cell.coordinate}"
                    cell_properties[cell_addr] = {
                        'value': str(cell_value),
                        'fill': self._get_fill_color(cell),
                        'font_color': self._get_font_color(cell),
                        'bold': cell.font.bold if cell.font else False
                    }

                data.append(row_data)

            sheets.append({
                'name': sheet.title,
                'max_row': sheet.max_row or 0,
                'max_col': sheet.max_column or 0,
                'data': data,
                'cell_properties': cell_properties,
                'has_merged': has_merged
            })

        return sheets

    def _get_fill_color(self, cell) -> str:
        """セルの背景色を取得"""
        try:
            if cell.fill and cell.fill.fgColor:
                color = cell.fill.fgColor
                if hasattr(color, 'rgb') and color.rgb:
                    return color.rgb
        except:
            pass
        return ""

    def _get_font_color(self, cell) -> str:
        """セルのフォント色を取得"""
        try:
            if cell.font and cell.font.color:
                color = cell.font.color
                if hasattr(color, 'rgb') and color.rgb:
                    return color.rgb
        except:
            pass
        return ""

    def _build_text(self, sheets: List[Dict[str, Any]]) -> str:
        """
        全シートのテキストを生成

        Returns:
            シート名とセルデータを結合したテキスト
        """
        result = []

        for sheet in sheets:
            result.append(f"\n[SHEET: {sheet['name']}]")
            for row in sheet['data']:
                result.append(" | ".join(row))
            result.append(f"[/SHEET]\n")

        return "\n".join(result)

    def _error_result(self, error_message: str) -> Dict[str, Any]:
        """エラー結果を返す"""
        return {
            'is_structured': False,
            'error': error_message,
            'text_with_tags': '',
            'structured_tables': [],
            'sheets': [],
            'tags': {},
            'media_elements': []
        }
