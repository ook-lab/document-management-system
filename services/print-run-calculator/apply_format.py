import json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def apply_strict_formatting(ws, format_json_path):
    with open(format_json_path, 'r', encoding='utf-8') as f:
        formats = json.load(f)

    # Set column widths
    for col_letter, width in formats.get('column_dimensions', {}).items():
        if width:
            ws.column_dimensions[col_letter].width = width

    # Set row heights
    for row_idx, height in formats.get('row_dimensions', {}).items():
        if height:
            ws.row_dimensions[int(row_idx)].height = height

    # Pre-define borders to save time (cache)
    border_styles = {}
    
    # Set cell formats
    for coord, cell_format in formats.get('cells', {}).items():
        cell = ws[coord]
        
        if 'font' in cell_format:
            f = cell_format['font']
            # Only set color if it's a valid hex string, openpyxl sometimes uses weird colors or Theme colors.
            # If color is '00000000', openpyxl might complain. Let's just pass the rgb string if it's hex-like.
            color = f.get('color')
            if color and color == '00000000':
                color = 'FF000000' # Default black
            
            try:
                cell.font = Font(
                    name=f.get('name'),
                    size=f.get('size'),
                    bold=f.get('bold'),
                    italic=f.get('italic'),
                    color=color
                )
            except ValueError:
                cell.font = Font(
                    name=f.get('name'),
                    size=f.get('size'),
                    bold=f.get('bold'),
                    italic=f.get('italic')
                )

        if 'fill' in cell_format:
            fill = cell_format['fill']
            if fill.get('patternType'):
                try:
                    cell.fill = PatternFill(
                        patternType=fill.get('patternType'),
                        fgColor=fill.get('fgColor')
                    )
                except ValueError:
                    pass

        if 'alignment' in cell_format:
            align = cell_format['alignment']
            cell.alignment = Alignment(
                horizontal=align.get('horizontal'),
                vertical=align.get('vertical'),
                wrap_text=align.get('wrap_text')
            )

        if 'border' in cell_format:
            b = cell_format['border']
            # create a hashable key for caching
            b_key = (b.get('top'), b.get('bottom'), b.get('left'), b.get('right'))
            if b_key not in border_styles:
                top = Side(style=b.get('top')) if b.get('top') else None
                bottom = Side(style=b.get('bottom')) if b.get('bottom') else None
                left = Side(style=b.get('left')) if b.get('left') else None
                right = Side(style=b.get('right')) if b.get('right') else None
                border_styles[b_key] = Border(top=top, bottom=bottom, left=left, right=right)
            
            cell.border = border_styles[b_key]
