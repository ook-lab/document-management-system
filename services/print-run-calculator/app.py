import os
import json
import uuid
import tempfile
from flask import Flask, render_template, request, send_file, redirect, url_for, flash, jsonify
import openpyxl
from apply_format import apply_strict_formatting

app = Flask(__name__)
app.secret_key = 'print_run_secret'

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'excel_templates')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'output')
CONFIG_FILE = os.path.join(DATA_DIR, 'series_config.json')
FORMAT_FILE = os.path.join(os.path.dirname(__file__), 'format_dump.json')
EXTRACTED_ELEMENTS_FILE = os.path.join(os.path.dirname(__file__), 'extracted_elements.json')

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(TEMPLATE_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

def load_config():
    if not os.path.exists(CONFIG_FILE):
        default_config = {"series": []}
        save_config(default_config)
        return default_config
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_config(config):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4, ensure_ascii=False)

@app.route('/')
def index():
    config = load_config()
    return render_template('index.html', series_list=config.get('series', []))

@app.route('/settings')
def settings():
    config = load_config()
    templates = [f for f in os.listdir(TEMPLATE_DIR) if f.endswith('.xlsm')] if os.path.exists(TEMPLATE_DIR) else []
    return render_template('settings.html', series_list=config.get('series', []), templates=templates)

@app.route('/settings/add', methods=['POST'])
def settings_add():
    config = load_config()
    
    # Load initial master elements if available
    elements = []
    if os.path.exists(EXTRACTED_ELEMENTS_FILE):
        with open(EXTRACTED_ELEMENTS_FILE, 'r', encoding='utf-8') as f:
            elements = json.load(f)
            
    new_series = {
        "id": str(uuid.uuid4()),
        "name": request.form.get('name'),
        "template_filename": request.form.get('template_filename'),
        "vba_code": request.form.get('vba_code'),
        "elements": elements
    }
    config.setdefault('series', []).append(new_series)
    save_config(config)
    return redirect(url_for('settings'))

@app.route('/settings/delete/<series_id>', methods=['POST'])
def settings_delete(series_id):
    config = load_config()
    config['series'] = [s for s in config.get('series', []) if s['id'] != series_id]
    save_config(config)
    return redirect(url_for('settings'))

@app.route('/settings/series/<series_id>')
def edit_series(series_id):
    config = load_config()
    series_data = next((s for s in config.get('series', []) if str(s.get('id')) == series_id), None)
    if not series_data:
        return "Series not found", 404
    return render_template('edit_series.html', series=series_data)

@app.route('/api/series/<series_id>/elements', methods=['POST'])
def save_series_elements(series_id):
    elements = request.json
    config = load_config()
    for s in config.get('series', []):
        if str(s.get('id')) == series_id:
            s['elements'] = elements
            save_config(config)
            return jsonify({"status": "ok"})
    return jsonify({"error": "not found"}), 404

@app.route('/api/series/<series_id>')
def get_series(series_id):
    config = load_config()
    for s in config.get('series', []):
        if str(s.get('id')) == series_id:
            return jsonify(s)
    return jsonify({}), 404

@app.route('/generate', methods=['POST'])
def generate():
    series_id = request.form.get('series_id')
    
    config = load_config()
    series_data = next((s for s in config.get('series', []) if str(s.get('id')) == series_id), None)
    
    if not series_data:
        return "Series not found", 400
        
    template_name = series_data.get('template_filename')
    if not template_name:
        return "No template assigned to this series", 400
        
    template_path = os.path.join(TEMPLATE_DIR, template_name)
    if not os.path.exists(template_path):
        return f"Template {template_name} not found", 404
        
    try:
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        ws = wb.active
        
        # Apply strict formatting explicitly from Python to prevent dependency on template layout
        if os.path.exists(FORMAT_FILE):
            apply_strict_formatting(ws, FORMAT_FILE)
            
        # Write ALL elements dynamically (Single Source of Truth)
        for el in series_data.get('elements', []):
            coord = el.get('coord')
            el_type = el.get('type')
            if not coord:
                continue
                
            if el_type == 'variable' or el_type == 'number':
                # User can override both variables and fixed numbers from the UI
                user_val = request.form.get(coord)
                
                # If user provided a value in the form, use it. Otherwise use master value.
                final_val = user_val if (user_val is not None and user_val != '') else el.get('value')
                
                try:
                    if isinstance(final_val, str) and '.' in final_val:
                        ws[coord] = float(final_val)
                    else:
                        ws[coord] = int(final_val)
                except (ValueError, TypeError):
                    ws[coord] = final_val
            elif el_type == 'formula':
                ws[coord] = el.get('value')
            elif el_type == 'text':
                ws[coord] = el.get('value')

        vba_code = series_data.get('vba_code')
        if vba_code:
            ws['AA1'] = vba_code

        book_title = request.form.get('出力ファイル名', 'output')
        output_filename = f"{book_title}.xlsm"
        
        temp_dir = tempfile.gettempdir()
        output_path = os.path.join(temp_dir, output_filename)
        wb.save(output_path)
        
        return send_file(output_path, as_attachment=True, download_name=output_filename)
    except Exception as e:
        return f"Error generating Excel file: {str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True, port=5007, host='0.0.0.0')
